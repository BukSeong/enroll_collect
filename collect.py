import sys
import os
import numpy as np
import pandas as pd
import json
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

# 兼容PyInstaller打包路径
if hasattr(sys, '_MEIPASS'):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(".")
logo_path = os.path.join(base_path, "logo_matrix.npy")
major_path = os.path.join(base_path, "major.json")

# 日志系统集成
import os
from datetime import datetime
from loguru import logger

# 日志目录和文件名
log_dir = os.path.join(os.getcwd(), "log")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

# 移除loguru默认的控制台输出
logger.remove()
# 添加文件日志输出
logger.add(log_file, encoding="utf-8", rotation="10 MB", retention="10 days", enqueue=True)

class MajorSelectDialog(QDialog):
    """意向专业选择弹窗"""
    def __init__(self, major_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择意向报考专业")
        self.setMinimumSize(600, 400)
        self.setWindowFlags(self.windowFlags() | Qt.WindowContextHelpButtonHint)
        self.major_data = major_data
        self.selected_college = None
        self.selected_major = None
        self.init_ui()

    def event(self, e):
        # 兼容PyQt5，QEvent.ContextHelp 可能不存在，直接用整数163判断
        # 163为QEvent.ContextHelp的实际值
        if e.type() == 163:
            QMessageBox.information(self, "操作说明", "1. 先选择学院，再选择专业，或直接输入‘学院-专业’进行搜索。\n2. 选择后点击‘确认’。\n3. 鼠标悬停可高亮选项。")
            return True
        return super().event(e)

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # 学院选择
        self.college_combo = QComboBox()
        self.college_combo.addItem("请选择学院")
        self.college_combo.addItems(list(self.major_data.keys()))
        self.college_combo.setStyleSheet("""
            QComboBox {
                padding: 12px 8px;
                font-size: 22px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 350px;
                min-height: 40px;
            }
            QComboBox:focus {
                border-color: #3498db;
            }
            QComboBox QAbstractItemView {
                selection-background-color: #ffe4b5;
                selection-color: #ca2b2f;
                background: #fff;
                font-size: 22px;
            }
        """)

        # 专业选择
        self.major_combo = QComboBox()
        self.major_combo.addItem("请选择专业")
        self.major_combo.setStyleSheet("""
            QComboBox {
                padding: 12px 8px;
                font-size: 22px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 350px;
                min-height: 40px;
            }
            QComboBox:focus {
                border-color: #3498db;
            }
            QComboBox QAbstractItemView {
                selection-background-color: #ffe4b5;
                selection-color: #ca2b2f;
                background: #fff;
                font-size: 22px;
            }
        """)
        self.major_combo.setEditable(True)
        self.major_combo.setInsertPolicy(QComboBox.NoInsert)

        # 搜索功能
        all_major_pairs = []
        for college, majors in self.major_data.items():
            for major in majors:
                all_major_pairs.append(f"{college}-{major}")
        completer = QCompleter(all_major_pairs)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setFilterMode(Qt.MatchContains)
        self.major_combo.setCompleter(completer)

        self.college_combo.currentTextChanged.connect(self.update_majors)
        self.major_combo.currentTextChanged.connect(self.major_select_entry)

        layout.addWidget(QLabel("学院："))
        layout.addWidget(self.college_combo)
        layout.addWidget(QLabel("专业："))
        layout.addWidget(self.major_combo)

        # 确认按钮
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(30)
        ok_btn = QPushButton("确认")
        ok_btn.setStyleSheet("font-size: 20px; padding: 10px 40px;")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("font-size: 20px; padding: 10px 40px;")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def update_majors(self, college):
        self.major_combo.clear()
        self.major_combo.addItem("请选择专业")
        if college in self.major_data:
            self.major_combo.addItems(self.major_data[college])

    def major_select_entry(self, text):
        if '-' in text:
            college, major = text.split('-', 1)
            if college in self.major_data and major in self.major_data[college]:
                self.college_combo.setCurrentText(college)
                self.major_combo.setCurrentText(major)

    def get_selection(self):
        college = self.college_combo.currentText()
        major = self.major_combo.currentText()
        if college == "请选择学院" or major == "请选择专业" or not major:
            return None, None
        return college, major


major_dict = {
    "管理学院":["大数据管理与应用","工商管理","工业工程","会计学（ACCA）"],
    "联合设计与创新学院":["工业设计","建筑学"],
    "经金学院":["金融工程","金融学","贸易经济（数字经济方向）","财政学","经济统计学","经济学","电子商务","国际经济与贸易"],
    "电气学院":["能源互联网工程","电气工程及其自动化"],
    "电信学部":["软件工程","计算机科学与技术","自动化","网络空间安全","电子科学与技术","物联网工程","微电子科学与工程","信息工程"],
    "钱学森学院":["储能科学与工程（新工科卓越计划）","计算机科学与技术（国家拔尖计划）","自动化（钱学森班本研一体）","能源与动力工程（钱学森班本研一体）","物理学（国家拔尖计划）","智能制造工程（钱学森班本研一体）","数学与应用数学（国家拔尖计划）","工程力学（国家拔尖计划）","少年班","基础医学（国家拔尖计划）","人工智能（新工科卓越计划）","临床医学（侯宗濂班本研一体）"],
    "马克思主义学院":["马克思主义理论"],
    "生命学院":["生物技术","生物医学工程"],
    "航天学院":["飞行器设计与工程","飞行器动力工程","工程力学"],
    "能动学院":["能源与动力工程（热流国际班）","能源与动力工程","环境工程","核工程与核技术","新能源科学与工程"],
    "化工学院":["化学工程与工艺","过程装备与控制工程"],
    "医学部":["药学","法医学","护理学","基础医学","口腔医学","医工学","制药工程","临床药学","临床医学（5+3一体化）","临床医学","预防医学"],
    "化学学院":["应用化学","化学"],
    "公管学院":["行政管理","劳动与社会保障"],
    "仪器学院":["测控技术与仪器"],
    "人文学院":["社会学","环境设计","汉语言文学","哲学","书法学"],
    "机械学院":["工业设计","车辆工程","机械工程（3D打印国际班）","机械工程","智能制造工程"],
    "物理学院":["材料物理","应用物理学","光电信息科学与工程"],
    "法学院":["法学","国际经贸规则"],
    "材料学院":["材料科学与工程"],
    "人居学院":["人居环境科学与技术"],
    "新闻学院":["网络与新媒体"],
    "数学学院":["统计学","数学与应用数学","信息与计算科学"],
    "外语学院":["英语（语言数据科学方向）","英语（英德方向）","英语","法语","日语（日英方向）","日语"]
}

class WelcomePage(QWidget):
    """欢迎页面"""
    def __init__(self, parent=None, logosize=512):
        super().__init__(parent)
        self.parent_window = parent
        self.logosize = logosize
        self.init_ui()
        
    def init_ui(self):
        # 设置背景色（非纯白色）
        self.setStyleSheet("background-color: #f5f5f5;")
        
        # 主布局
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignCenter)
        
        # 校徽显示
        self.logo_label = QLabel()
        self.logo_label.setAlignment(Qt.AlignCenter)
        self.display_logo()
        
        # 欢迎文本
        welcome_text = QLabel("欢迎使用西安交通大学-青岛五十八中意向生收集系统！")
        welcome_text.setAlignment(Qt.AlignCenter)
        welcome_text.setStyleSheet("font-size: 38px; font-weight: bold; color: #333; margin: 40px;")

        # 作者信息
        author_text = QLabel("Author: BukSeong(https://github.com/BukSeong). Please indicate the source when reposting.")
        author_text.setAlignment(Qt.AlignCenter)
        author_text.setStyleSheet("font-size: 22px; color: #666; margin: 20px;")
        
        # 按钮布局
        button_layout = QHBoxLayout()
        button_layout.setAlignment(Qt.AlignCenter)
        button_layout.setSpacing(30)
        
        # 确认使用按钮
        confirm_btn = QPushButton("确认使用")
        confirm_btn.setStyleSheet("""
            QPushButton {
                background-color: #ca2b2f;
                color: white;
                font-size: 28px;
                font-weight: bold;
                padding: 25px 70px;
                border-radius: 12px;
                border: none;
            }
            QPushButton:hover {
                background-color: #e74c3c;
            }
            QPushButton:pressed {
                background-color: #c0392b;
            }
        """)
        confirm_btn.clicked.connect(self.parent_window.show_form_page)
        
        # 检查导出按钮
        check_btn = QPushButton("检查导出")
        check_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-size: 28px;
                font-weight: bold;
                padding: 25px 70px;
                border-radius: 12px;
                border: none;
            }
            QPushButton:hover {
                background-color: #5dade2;
            }
            QPushButton:pressed {
                background-color: #2980b9;
            }
        """)
        check_btn.clicked.connect(self.check_export)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-size: 28px;
                font-weight: bold;
                padding: 25px 70px;
                border-radius: 12px;
                border: none;
            }
            QPushButton:hover {
                background-color: #bdc3c7;
            }
            QPushButton:pressed {
                background-color: #7f8c8d;
            }
        """)
        close_btn.clicked.connect(self.parent_window.close)
        
        button_layout.addWidget(confirm_btn)
        button_layout.addWidget(check_btn)
        button_layout.addWidget(close_btn)
        
        # 添加到主布局
        main_layout.addWidget(self.logo_label)
        main_layout.addWidget(welcome_text)
        main_layout.addWidget(author_text)
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
    
    def display_logo(self):
        """显示校徽点阵"""
        try:
            matrix = np.load(logo_path)
            size = self.logosize if hasattr(self, 'logosize') else 256
            pixmap = QPixmap(size, size)
            pixmap.fill(QColor(245, 245, 245))

            painter = QPainter(pixmap)
            red_color = QColor(202, 43, 47)

            scale_x = size / matrix.shape[1]
            scale_y = size / matrix.shape[0]

            for i in range(matrix.shape[0]):
                for j in range(matrix.shape[1]):
                    if matrix[i, j] == 1:
                        x = int(j * scale_x)
                        y = int(i * scale_y)
                        painter.setPen(red_color)
                        painter.setBrush(red_color)
                        painter.drawRect(x, y, max(1, int(scale_x)), max(1, int(scale_y)))

            painter.end()
            self.logo_label.setPixmap(pixmap)
        except Exception as e:
            logger.error(f"加载校徽失败: {e}")
            self.logo_label.setText("校徽加载失败")
            self.logo_label.setStyleSheet("font-size: 20px; color: red;")
    
    def check_export(self):
        """检查导出的.xlsx文件"""
        try:
            if not os.path.exists('res.xlsx'):
                QMessageBox.information(self, "检查结果", "尚未创建导出文件，暂无学生信息。")
                return
            
            # 读取Excel文件
            df = pd.read_excel('res.xlsx')
            
            # 检查信息
            total_students = len(df)
            incomplete_info = 0
            
            # 检查每一行的完整性（备注不计入不完整）
            for index, row in df.iterrows():
                row_check = row.drop(labels=['备注']) if '备注' in row.index else row
                if row_check.isnull().any():
                    incomplete_info += 1
            
            # 弹窗提示
            msg = f"共有 {total_students} 位学生信息。\n"
            if incomplete_info > 0:
                msg += f"其中有 {incomplete_info} 位学生信息不完整。"
            else:
                msg += "所有学生信息完整。"
            
            QMessageBox.information(self, "检查结果", msg)
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"检查导出文件时出错: {e}")


class FormPage(QWidget):
    """表单页面"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        # 直接使用major_dict
        self.major_data = major_dict
        self.init_ui()
    
    def load_major_data(self):
        """加载专业数据"""
        try:
            with open(major_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"加载专业数据失败: {e}")
            return {}
    
    def init_ui(self):
        # 设置背景色
        self.setStyleSheet("background-color: #f5f5f5;")
        
        # 主布局
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        
        # 标题
        title = QLabel("意向生信息收集表")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #ca2b2f; margin: 20px;")
        
        # 表单布局
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setFormAlignment(Qt.AlignCenter)
        form_layout.setSpacing(20)
        
        # 1. 姓名
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("请输入姓名")
        self.name_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 2. 所在中学
        self.school_input = QLineEdit("山东省青岛第五十八中学")
        self.school_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 3. 联系电话
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("请输入11位手机号")
        # 不再限制最大长度，也不设置IntValidator，允许任意输入
        self.phone_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 4. 选考科目
        self.subject_widget = self.create_subject_widget()
        
        # 5. 意向报考专业（按钮弹窗选择）
        self.selected_college = None
        self.selected_major = None
        self.major_btn = QPushButton("选择意向报考专业")
        self.major_btn.setStyleSheet("font-size: 22px; padding: 12px 40px; background-color: #f5f5f5; border: 2px solid #ca2b2f; border-radius: 8px; color: #ca2b2f;")
        self.major_btn.clicked.connect(self.open_major_dialog)
        
        # 6. 最近一次考试分数
        self.score_input = QLineEdit()
        self.score_input.setPlaceholderText("请输入分数")
        self.score_input.setValidator(QDoubleValidator())
        self.score_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 7. 总分数
        self.total_score_input = QLineEdit("750")
        self.total_score_input.setValidator(QDoubleValidator())
        self.total_score_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 8. 最近一次年级排名
        self.rank_input = QLineEdit()
        self.rank_input.setPlaceholderText("请输入排名")
        self.rank_input.setValidator(QIntValidator())
        self.rank_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)

        # 10. 备注
        self.remark_input = QLineEdit()
        self.remark_input.setPlaceholderText("请输入备注")
        self.remark_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 9. 参加排名人数
        self.total_rank_input = QLineEdit("1000")
        self.total_rank_input.setValidator(QIntValidator())
        self.total_rank_input.setStyleSheet("""
            QLineEdit {
                padding: 15px 10px;
                font-size: 24px;
                border: 2px solid #ddd;
                border-radius: 5px;
                min-width: 450px;
                min-height: 56px;
                line-height: 32px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        
        # 添加到表单
        form_layout.addRow("姓名：", self.name_input)
        form_layout.addRow("所在中学：", self.school_input)
        form_layout.addRow("联系电话：", self.phone_input)
        form_layout.addRow("选考科目：", self.subject_widget)
        form_layout.addRow("意向报考专业：", self.major_btn)
        form_layout.addRow("最近一次考试分数：", self.score_input)
        form_layout.addRow("总分数：", self.total_score_input)
        form_layout.addRow("最近一次年级排名：", self.rank_input)
        form_layout.addRow("参加排名人数：", self.total_rank_input)
        form_layout.addRow("备注：", self.remark_input)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.setAlignment(Qt.AlignCenter)
        button_layout.setSpacing(30)
        
        submit_btn = QPushButton("确认并导出")
        submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                font-size: 18px;
                font-weight: bold;
                padding: 15px 40px;
                border-radius: 8px;
                border: none;
                margin-top: 30px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
            QPushButton:pressed {
                background-color: #229954;
            }
        """)
        submit_btn.clicked.connect(self.submit_form)
        
        back_btn = QPushButton("返回")
        back_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-size: 18px;
                font-weight: bold;
                padding: 15px 40px;
                border-radius: 8px;
                border: none;
                margin-top: 30px;
            }
            QPushButton:hover {
                background-color: #bdc3c7;
            }
            QPushButton:pressed {
                background-color: #7f8c8d;
            }
        """)
        back_btn.clicked.connect(self.parent_window.show_welcome_page)
        
        button_layout.addWidget(submit_btn)
        button_layout.addWidget(back_btn)
        
        # 添加到主布局
        main_layout.addWidget(title)
        main_layout.addLayout(form_layout)
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
    
    def create_subject_widget(self):
        """创建选考科目选择控件"""
        widget = QWidget()
        layout = QHBoxLayout()
        layout.setSpacing(15)
        
        self.subject_checkboxes = {}
        subjects = ["物理", "化学", "生物", "政治", "历史", "地理"]
        
        for subject in subjects:
            checkbox = QCheckBox(subject)
            checkbox.setStyleSheet("""
                QCheckBox {
                    font-size: 16px;
                    spacing: 8px;
                }
                QCheckBox::indicator {
                    width: 20px;
                    height: 20px;
                }
            """)
            self.subject_checkboxes[subject] = checkbox
            layout.addWidget(checkbox)
        
        widget.setLayout(layout)
        return widget
    
    def open_major_dialog(self):
        dialog = MajorSelectDialog(self.major_data, self)
        if dialog.exec_() == QDialog.Accepted:
            college, major = dialog.get_selection()
            if college and major:
                self.selected_college = college
                self.selected_major = major
                self.major_btn.setText(f"{college} - {major}")
            else:
                self.selected_college = None
                self.selected_major = None
                self.major_btn.setText("选择意向报考专业")
    
    def update_majors(self, college):
        """更新专业下拉框"""
        self.major_combo.clear()
        self.major_combo.addItem("请选择专业")
        if college in self.major_data:
            self.major_combo.addItems(self.major_data[college])
    
    def get_selected_subjects(self):
        """获取选中的科目"""
        selected = []
        for subject, checkbox in self.subject_checkboxes.items():
            if checkbox.isChecked():
                selected.append(subject)
        return selected
    
    def validate_form(self):
        """验证表单数据"""
        # 1. 姓名不能为空
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "验证错误", "请输入姓名！")
            return False
        
        # 2. 联系电话必须是11位数字
        phone = self.phone_input.text().strip()
        if len(phone) != 11 or not phone.isdigit():
            QMessageBox.warning(self, "验证错误", "请输入正确的11位手机号！")
            return False
        
        # 3. 选考科目必须六选三
        selected_subjects = self.get_selected_subjects()
        if len(selected_subjects) != 3:
            QMessageBox.warning(self, "验证错误", "请选择3门选考科目！")
            return False
        
        # 4. 意向报考专业必须选择
        if not self.selected_college or not self.selected_major:
            QMessageBox.warning(self, "验证错误", "请选择意向报考专业！")
            return False
        
        # 5. 最近一次考试分数不能为空
        if not self.score_input.text().strip():
            QMessageBox.warning(self, "验证错误", "请输入最近一次考试分数！")
            return False
        
        # 6. 总分数不能为空
        if not self.total_score_input.text().strip():
            QMessageBox.warning(self, "验证错误", "请输入总分数！")
            return False
        
        # 7. 最近一次年级排名不能为空
        if not self.rank_input.text().strip():
            QMessageBox.warning(self, "验证错误", "请输入最近一次年级排名！")
            return False
        
        # 8. 参加排名人数不能为空
        if not self.total_rank_input.text().strip():
            QMessageBox.warning(self, "验证错误", "请输入参加排名人数！")
            return False

        # 备注可为空，无需校验
        
        return True
    
    def submit_form(self):
        """提交表单"""
        if not self.validate_form():
            return
        try:
            # 收集数据
            student_data = {
                '姓名': self.name_input.text().strip(),
                '所在中学': self.school_input.text().strip(),
                '联系电话': str(self.phone_input.text().strip()),  # 强制为字符串
                '选考科目': '、'.join(self.get_selected_subjects()),
                '意向学院': self.selected_college,
                '意向专业': self.selected_major,
                '最近一次考试分数': float(self.score_input.text()),
                '总分数': float(self.total_score_input.text()),
                '最近一次年级排名': int(self.rank_input.text()),
                '参加排名人数': int(self.total_rank_input.text()),
                '备注': self.remark_input.text().strip()
            }
            # 导出到Excel
            self.export_to_excel(student_data)
            # 清空表单（为下一位学生准备）
            self.clear_form()
            # 显示成功页面
            self.parent_window.show_success_page()
        except Exception as e:
            QMessageBox.critical(self, "导出错误", f"导出数据时出错: {e}")
    
    def export_to_excel(self, student_data):
        """导出数据到Excel，手机号强制为文本格式"""
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import numbers
        import pandas as pd
        import os
        df_new = pd.DataFrame([student_data])
        if not os.path.exists('res.xlsx'):
            # 新建文件，手机号列设置为文本格式
            with pd.ExcelWriter('res.xlsx', engine='openpyxl') as writer:
                df_new.to_excel(writer, index=True, index_label='序号')
                ws = writer.book.active
                for cell in ws[1]:
                    if cell.value == '联系电话':
                        col_letter = cell.column_letter
                        for row in ws.iter_rows(min_row=2, min_col=cell.col_idx, max_col=cell.col_idx):
                            for c in row:
                                c.number_format = '@'
        else:
            df_existing = pd.read_excel('res.xlsx', index_col=0, dtype={'联系电话':str})
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            df_combined.index = df_combined.index + 1
            with pd.ExcelWriter('res.xlsx', engine='openpyxl') as writer:
                df_combined.to_excel(writer, index=True, index_label='序号')
                ws = writer.book.active
                for cell in ws[1]:
                    if cell.value == '联系电话':
                        col_letter = cell.column_letter
                        for row in ws.iter_rows(min_row=2, min_col=cell.col_idx, max_col=cell.col_idx):
                            for c in row:
                                c.number_format = '@'
    
    def clear_form(self):
        """清空表单"""
        self.name_input.clear()
        self.school_input.setText("山东省青岛第五十八中学")
        self.phone_input.clear()
        
        # 清空选考科目
        for checkbox in self.subject_checkboxes.values():
            checkbox.setChecked(False)
        
        # 重置专业选择
        self.selected_college = None
        self.selected_major = None
        self.major_btn.setText("选择意向报考专业")
        
        self.score_input.clear()
        self.total_score_input.setText("750")
        self.rank_input.clear()
        self.total_rank_input.setText("1000")
        self.remark_input.clear()



class SuccessPage(QWidget):
    """成功页面"""
    def __init__(self, parent=None, logosize=256):
        super().__init__(parent)
        self.parent_window = parent
        self.logosize = logosize
        self.init_ui()

    def init_ui(self):
        # 设置背景色
        self.setStyleSheet("background-color: #f5f5f5;")
        # 主布局
        self.main_layout = QVBoxLayout()
        self.main_layout.setAlignment(Qt.AlignCenter)

        # 成功信息
        self.success_text = QLabel()
        self.success_text.setAlignment(Qt.AlignCenter)
        self.success_text.setStyleSheet("font-size: 28px; font-weight: bold; color: #2ecc71; margin: 30px;")

        # 随机祝福
        self.wish_text = QLabel()
        self.wish_text.setAlignment(Qt.AlignCenter)
        self.wish_text.setStyleSheet("font-size: 24px; color: #333; margin: 20px;")

        # 校徽显示
        self.logo_label = QLabel()
        self.logo_label.setAlignment(Qt.AlignCenter)
        self.display_logo()

        # 返回按钮
        back_btn = QPushButton("返回")
        back_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-size: 18px;
                font-weight: bold;
                padding: 15px 40px;
                border-radius: 8px;
                border: none;
                margin-top: 30px;
            }
            QPushButton:hover {
                background-color: #5dade2;
            }
            QPushButton:pressed {
                background-color: #2980b9;
            }
        """)
        back_btn.clicked.connect(self.parent_window.show_welcome_page)

        self.main_layout.addWidget(self.success_text)
        self.main_layout.addWidget(self.wish_text)
        self.main_layout.addWidget(self.logo_label)
        self.main_layout.addWidget(back_btn)
        self.setLayout(self.main_layout)

    def refresh(self):
        import random
        greetings = [
            "春风得意马蹄疾，一日看尽长安花。",
            "愿你金榜题名，前程似锦！",
            "星光不负赶路人，时光不负有心人。",
            "愿你不负韶华，圆梦交大！",
            "高考顺利，梦想成真！",
            "愿你前路光明，未来可期！",
            "一举夺魁，鹏程万里！",
            "愿你旗开得胜，马到成功！",
            "愿你心想事成，前程似锦！",
            "愿你以梦为马，不负韶华！"
        ]
        student_count = self.get_student_count() - 1 if self.get_student_count() != 1 else 1 # 减1因为刚添加完
        self.success_text.setText(f"恭喜你成为第{student_count}位意向生！")
        wish = random.choice(greetings)
        self.wish_text.setText(wish)

    def get_student_count(self):
        try:
            if os.path.exists('res.xlsx'):
                df = pd.read_excel('res.xlsx', index_col=0)
                if '序号' in df.index.name or df.index.name == '序号':
                    if len(df.index) > 0:
                        max_idx = max(df.index)
                        return int(max_idx) + 1
                    else:
                        return 1
                else:
                    return len(df) + 1
            else:
                return 1
        except Exception as e:
            return 1

    def display_logo(self):
        try:
            matrix = np.load(logo_path)
            size = self.logosize if hasattr(self, 'logosize') else 256
            pixmap = QPixmap(size, size)
            pixmap.fill(QColor(245, 245, 245))

            painter = QPainter(pixmap)
            red_color = QColor(202, 43, 47)

            scale_x = size / matrix.shape[1]
            scale_y = size / matrix.shape[0]

            for i in range(matrix.shape[0]):
                for j in range(matrix.shape[1]):
                    if matrix[i, j] == 1:
                        x = int(j * scale_x)
                        y = int(i * scale_y)
                        painter.setPen(red_color)
                        painter.setBrush(red_color)
                        painter.drawRect(x, y, max(1, int(scale_x)), max(1, int(scale_y)))

            painter.end()
            self.logo_label.setPixmap(pixmap)
        except Exception as e:
            logger.error(f"加载校徽失败: {e}")
            self.logo_label.setText("校徽加载失败")
            self.logo_label.setStyleSheet("font-size: 20px; color: red;")


class MainWindow(QMainWindow):
    """主窗口"""
    def __init__(self, logosize=256):
        super().__init__()
        self.logosize = logosize
        self.init_ui()
    
    def init_ui(self):
        # 设置窗口标题
        self.setWindowTitle("西安交通大学-青岛五十八中意向生收集系统")
        
        # 设置窗口大小（8:5比例）
        self.resize(1920, 1200)  # 默认更大
        self.setMinimumSize(1200, 700)  # 最小更大，避免缩放问题
        
        # 设置中心部件
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        # 使用堆叠布局
        self.stacked_layout = QStackedLayout()
        self.central_widget.setLayout(self.stacked_layout)
        
        # 创建各个页面，传递logosize
        self.welcome_page = WelcomePage(self, logosize=self.logosize)
        self.form_page = FormPage(self)
        self.success_page = SuccessPage(self, logosize=self.logosize)
        
        # 添加到堆叠布局
        self.stacked_layout.addWidget(self.welcome_page)
        self.stacked_layout.addWidget(self.form_page)
        self.stacked_layout.addWidget(self.success_page)
        
        # 显示欢迎页面
        self.show_welcome_page()
    
    def show_welcome_page(self):
        """显示欢迎页面"""
        self.stacked_layout.setCurrentWidget(self.welcome_page)
    
    def show_form_page(self):
        """显示表单页面"""
        self.stacked_layout.setCurrentWidget(self.form_page)
    
    def show_success_page(self):
        """显示成功页面"""
        self.success_page.refresh()
        self.stacked_layout.setCurrentWidget(self.success_page)


def main():
    app = QApplication(sys.argv)
    
    # 设置应用程序图标和样式
    app.setStyle('Fusion')
    
    # 设置中文字体（放大）
    font = QFont()
    font.setFamily("STZhongsong")
    font.setPointSize(18)  # 更大字体
    app.setFont(font)
    
    # 创建并显示主窗口
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    import os
    
    # 确保所需文件存在
    if not os.path.exists('logo_matrix.npy'):
        logger.warning("logo_matrix.npy 文件不存在！")
        logger.warning("请先运行 base01.py 生成校徽点阵文件。")
        # 创建默认的128x128点阵（中间一个红色方块作为示例）
        default_matrix = np.zeros((128, 128), dtype=np.uint8)
        default_matrix[50:78, 50:78] = 1
        np.save('logo_matrix.npy', default_matrix)
        logger.info("已创建默认校徽点阵文件。")

    if not os.path.exists('major.json'):
        logger.error("major.json 文件不存在！")
        logger.error("请确保专业数据文件存在。")
        sys.exit(1)

    main()
